"""
Script de ejemplo para el tipo "Aire" del módulo Cliente X.
Reemplazar la lógica interna por la real (lectura de columnas específicas,
mapeo a esquema de BigQuery, etc.) — la firma run(file_path, params, log)
es lo único que el resto del sistema necesita respetar.
"""
import csv
import os
from openpyxl import load_workbook


def run(file_path: str, params: dict, log) -> dict:
    log("Leyendo archivo .xlsx…")
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    log("Convirtiendo XLSX → CSV…")
    csv_path = os.path.splitext(file_path)[0] + ".csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        rows = 0
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)
            rows += 1

    log(f"{rows} filas convertidas.")
    log("[PRUEBA] Simulando carga a BigQuery (credenciales reales pendientes)…")
    table = f"cliente_x.aire_{params.get('periodo', 'sin_fecha')}"

    return {"bigquery_table": table, "filas": rows, "csv_path": csv_path}
